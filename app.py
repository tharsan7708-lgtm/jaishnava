from flask import Flask, render_template, request, redirect, session, url_for, flash, send_file
from flask_mysqldb import MySQL
from datetime import datetime, date
import MySQLdb.cursors
from flask import Response


import tempfile
import os
from werkzeug.utils import secure_filename

from flask import send_file
import csv
from io import StringIO

app = Flask(__name__)
app.secret_key = 'your_secret_key'

UPLOAD_FOLDER = 'static/profile/'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


# MySQL Config
app.config['MYSQL_HOST'] = 'jd-jaishnavadesigners.l.aivencloud.com'
app.config['MYSQL_USER'] = 'avnadmin'
app.config['MYSQL_PASSWORD'] = 'AVNS_MdgTDtpwxMHwJs4GSWk'
app.config['MYSQL_DB'] = 'jaishnava'
app.config['MYSQL_PORT']=27602

mysql = MySQL(app)

@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        mobile = request.form['mobile']
        secret = request.form['secret']
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute("SELECT * FROM users WHERE mobile=%s AND secret=%s", (mobile, secret))
        user = cursor.fetchone()
        if user:
            session['loggedin'] = True
            session['mobile'] = user['mobile']
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid Login')
    return render_template('login.html')

@app.route('/dashboard')
def dashboard():
    if not session.get('loggedin'):
        return redirect(url_for('login'))
    
    cursor = mysql.connection.cursor()
    cursor.execute("SELECT COUNT(*) FROM employees")
    employee_count = cursor.fetchone()[0]
    cursor.close()

    return render_template('dashboard.html', employee_count=employee_count)


import os
from datetime import date
from flask import request, render_template, redirect, url_for, flash
from werkzeug.utils import secure_filename

# Set upload folder path
UPLOAD_FOLDER = 'static/profile/'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/add_employee', methods=['GET', 'POST'])
def add_employee():
    if request.method == 'POST':
        emp_id = request.form['emp_id']
        name = request.form['name']
        gender=request.form['gender']
        job = request.form['job']
        salary = request.form['salary']
        dof = request.form['dof']
        profile_pic = request.files['profile_pic']

        # Handle profile picture upload and renaming
        if profile_pic and profile_pic.filename != '':
            original_filename = secure_filename(profile_pic.filename)
            ext = os.path.splitext(original_filename)[1]  # Get file extension
            safe_name = name.replace(" ", "_")  # Remove spaces from name
            new_filename = f"{emp_id}_{safe_name}{ext}"
            save_path = os.path.join(app.config['UPLOAD_FOLDER'], new_filename)
            profile_pic.save(save_path)
        else:
            safe_name = name.replace(" ", "_") 
            new_filename = f"{emp_id}_{safe_name}"+ ".png"

        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

        # Check for duplicate emp_id
        cursor.execute("SELECT * FROM employees WHERE emp_id = %s", (emp_id,))
        existing = cursor.fetchone()
        if existing:
            flash("Employee ID already exists")
            return redirect(url_for('add_employee'))

        # Insert into employees table
        cursor.execute(
            "INSERT INTO employees (emp_id, name,gender, job, salary_per_day, dof, profile_pic) VALUES (%s, %s,%s, %s, %s, %s, %s)",
            (emp_id, name,gender, job, salary, dof, new_filename)
        )
        mysql.connection.commit()

        # Create attendance table for this employee
        table_name = f"attendance_{emp_id}"
        cursor.execute(f"""
            CREATE TABLE IF NOT EXISTS {table_name} (
                id INT AUTO_INCREMENT PRIMARY KEY,
                date DATE NOT NULL UNIQUE,
                status ENUM('full', 'half', 'absent') NOT NULL
            )
        """)
        mysql.connection.commit()

        return redirect(url_for('dashboard'))

    return render_template('add_employee.html')

@app.route('/attendance', methods=['GET', 'POST'])
def attendance():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("SELECT name FROM employees")
    employees = cursor.fetchall()
    if request.method == 'POST':
        name = request.form['name']
        att_date = request.form['date']
        status = request.form['status']
        today = date.today()
        att_date_obj = datetime.strptime(att_date, "%Y-%m-%d").date()
        
        cursor.execute("SELECT emp_id FROM employees WHERE name=%s",(name,))
        a=cursor.fetchone()['emp_id'] or 0
        
        if att_date_obj > today:
            flash("Cannot mark attendance for future date")
            return redirect(url_for('attendance'))
        tablename = f"attendance_{a}"
        try:
            q=f"INSERT INTO {tablename} (date,status) VALUES (%s,%s)"
            cursor.execute(q,(att_date,status))

            mysql.connection.commit()
        except:
            flash("Attendance already marked for this date")
    return render_template('attendance.html', employees=employees)

@app.route('/salary', methods=['GET', 'POST'])
def salary():
    if not session.get('loggedin'):
        return redirect(url_for('login'))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("SELECT * FROM employees")
    employees = cursor.fetchall()

    current_date = date.today()
    selected_month = current_date.month
    selected_year = current_date.year

    if request.method == 'POST':
        selected_month = int(request.form['month'])
        selected_year = int(request.form['year'])

    salary_data = []
    total_salary_sum = 0
    total_bonus_sum = 0
    total_advance_sum = 0
    total_pending_sum = 0
    outstanding_amount = 0

    for emp in employees:
        table_name = f"attendance_{emp['emp_id']}"

        # Calculate effective present days
        cursor.execute(f"""
            SELECT 
                SUM(CASE 
                        WHEN status = 'full' THEN 1
                        WHEN status = 'half' THEN 0.5
                        ELSE 0 
                    END) AS effective_days
            FROM {table_name}
            WHERE MONTH(date) = %s AND YEAR(date) = %s
        """, (selected_month, selected_year))
        
        result = cursor.fetchone()
        effective_days = result['effective_days'] or 0

        total_salary = round(effective_days * emp['salary_per_day'], 2)

        # Get bonus amount
        cursor.execute("""
            SELECT SUM(amount) AS bonus 
            FROM bonuses 
            WHERE employee_name = %s AND MONTH(date) = %s AND YEAR(date) = %s
        """, (emp['name'], selected_month, selected_year))
        bonus = cursor.fetchone()['bonus'] or 0

        # Get advance amount
        cursor.execute("""
            SELECT SUM(amount) AS advance 
            FROM advance 
            WHERE employee_name = %s AND MONTH(date) = %s AND YEAR(date) = %s
        """, (emp['name'], selected_month, selected_year))
        advance = cursor.fetchone()['advance'] or 0

        pending = round(total_salary + bonus - advance, 2)

        # Accumulate totals
        total_salary_sum += total_salary
        total_bonus_sum += bonus
        total_advance_sum += advance

        if pending < 0:
            outstanding_amount += abs(pending)
        elif pending > 0:
            total_pending_sum += pending

        salary_data.append({
            'emp_id': emp['emp_id'],
            'name': emp['name'],
            'present_days': effective_days,
            'salary': total_salary,
            'bonus': bonus,
            'advance': advance,
            'pending': pending
        })

    cursor.close()

    return render_template('salary.html',
                           salary_data=salary_data,
                           total_salary=round(total_salary_sum+total_bonus_sum, 2),
                           total_bonus=round(total_bonus_sum, 2),
                           total_advance=round(total_advance_sum, 2),
                           total_pending=round(total_pending_sum, 2),
                           selected_month=selected_month,
                           selected_year=selected_year,
                           current_year=date.today().year,
                           outstanding_amount=round(outstanding_amount, 2))

@app.route('/advance', methods=['GET', 'POST'])
def advance():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # Fetch employee names for dropdown
    cursor.execute("SELECT name FROM employees")
    employees = cursor.fetchall()

    if request.method == 'POST':
        name = request.form['name']
        amount = request.form['amount']
        advance_date = request.form['date']  # Date from the form (string)

        # Optional: validate or convert date string
        try:
            date_obj = datetime.strptime(advance_date, "%Y-%m-%d").date()
        except ValueError:
            flash("Invalid date format")
            return redirect(url_for('advance'))
        

        cursor.execute("SELECT emp_id FROM employees WHERE name=%s",(name,))
        a=cursor.fetchone()['emp_id'] or 0
        

        # Insert advance payment record
        cursor.execute(
            "INSERT INTO advance (emp_id,employee_name, amount, date) VALUES (%s,%s, %s, %s)",
            (a,name, amount, date_obj)
        )
        mysql.connection.commit()
        return redirect(url_for('dashboard'))

    return render_template('advance.html', employees=employees)


@app.route('/bonus', methods=['GET', 'POST'])
def bonus():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # Fetch employee names for dropdown
    cursor.execute("SELECT name FROM employees")
    employees = cursor.fetchall()

    if request.method == 'POST':
        name = request.form['name']
        amount = request.form['amount']
        bonus_date = request.form['date']  # Date from the form (string)

        # Optional: validate or convert date string
        try:
            date_obj = datetime.strptime(bonus_date, "%Y-%m-%d").date()
        except ValueError:
            flash("Invalid date format")
            return redirect(url_for('bonus'))
        
        cursor.execute("SELECT emp_id FROM employees WHERE name=%s",(name,))
        a=cursor.fetchone()['emp_id'] or 0

        # Insert advance payment record
        cursor.execute(
            "INSERT INTO bonuses (emp_id,employee_name, amount, date) VALUES (%s,%s, %s, %s)",
            (a,name, amount, date_obj)
        )
        mysql.connection.commit()
        return redirect(url_for('dashboard'))

    return render_template('bonus.html', employees=employees)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

from flask import Flask, render_template, request, redirect, url_for, flash
from flask_mysqldb import MySQL
from datetime import datetime, date
import os
import MySQLdb.cursors

@app.route('/view_employees', methods=['GET', 'POST'])
def view_employees():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Handle employee addition
    if request.method == 'POST':
        emp_id = request.form['emp_id']
        name = request.form['name']
        job = request.form['job']
        salary = request.form['salary']
        dof = request.form['dof']
        profile_pic = request.files['profile_pic']

        # Rename profile pic to empid_name.extension
        filename = f"{emp_id}_{name.replace(' ', '_')}.{profile_pic.filename.rsplit('.', 1)[-1]}"
        filepath = os.path.join('static/profile',filename)
        profile_pic.save(filepath)

        # Check for duplicate emp_id
        cursor.execute("SELECT * FROM employees WHERE emp_id = %s", (emp_id,))
        if cursor.fetchone():
            flash("Employee ID already exists")
            return redirect(url_for('view_employees'))

        # Insert new employee
        cursor.execute("""INSERT INTO employees (emp_id, name, job, salary_per_day, dof, profile_pic)
                          VALUES (%s, %s, %s, %s, %s, %s)""",
                       (emp_id, name, job, salary, dof, filename))
        mysql.connection.commit()

        # Create employee-specific attendance table
        cursor.execute(f"""CREATE TABLE IF NOT EXISTS attendance_{name.replace(' ', '_')} (
                            id INT AUTO_INCREMENT PRIMARY KEY,
                            date DATE NOT NULL UNIQUE,
                            status ENUM('full', 'half', 'absent') NOT NULL)""")
        mysql.connection.commit()
        return redirect(url_for('view_employees'))

    # Fetch employees
    cursor.execute("SELECT * FROM employees")
    employees = cursor.fetchall()

    # Calculate experience without using dateutil
    for emp in employees:
        if emp['dof']:
            joining_date = emp['dof'] if isinstance(emp['dof'], date) else datetime.strptime(str(emp['dof']), "%Y-%m-%d").date()
            today = date.today()
            years = today.year - joining_date.year
            months = today.month - joining_date.month

            if months < 0:
                years -= 1
                months += 12

            emp['experience'] = f"{years}Y {months}M"
        else:
            emp['experience'] = "N/A"

    return render_template('view_employees.html', employees=employees)

@app.route('/edit_employee/<int:emp_id>', methods=['GET', 'POST'])
def edit_employee(emp_id):
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Fetch employee by ID
    cursor.execute("SELECT * FROM employees WHERE emp_id = %s", (emp_id,))
    employee = cursor.fetchone()

    if not employee:
        flash("Employee not found", "error")
        return redirect(url_for('view_employees'))

    if request.method == 'POST':
        # Get new values from form
        new_job = request.form['job']
        new_salary = request.form['salary']

        # Update in DB
        cursor.execute("UPDATE employees SET job = %s, salary_per_day = %s WHERE emp_id = %s",
                       (new_job, new_salary, emp_id))
        mysql.connection.commit()
        flash("Employee details updated successfully", "success")
        return redirect(url_for('view_employees'))

    return render_template('edit_employee.html', employee=employee)


@app.route('/delete_employee/<emp_id>', methods=['GET', 'POST'])
def delete_employee(emp_id):
    cursor = mysql.connection.cursor()
    if request.method == 'POST':
        cursor.execute("DELETE FROM employees WHERE emp_id = %s", (emp_id,))
        mysql.connection.commit()
        
        flash('Employee deleted successfully!', 'success')
        return redirect(url_for('view_employees'))
    else:
        cursor.execute("SELECT * FROM employees WHERE emp_id = %s", (emp_id,))
        emp = cursor.fetchone()
        return render_template('delete_employee.html', emp=emp)
    
from flask import request, render_template
from datetime import datetime

@app.route("/calendar", methods=["GET", "POST"])
def calendar():
    if request.method == "POST":
        month = int(request.form.get("month", datetime.now().month))
        year = int(request.form.get("year", datetime.now().year))
    else:
        month = datetime.now().month
        year = datetime.now().year

    cursor = mysql.connection.cursor()
    cursor.execute("SELECT emp_id, name, profile_pic FROM employees")
    employees = [dict(id=row[0], name=row[1], profile_picture=row[2]) for row in cursor.fetchall()]
    
    return render_template(
        "calendar.html",
        employees=employees,
        selected_month=month,
        selected_year=year,
        current_year=datetime.now().year
    )

    



from flask import send_file
import csv, tempfile, os
from datetime import datetime

@app.route('/download_salary_csv/<int:month>/<int:year>')
def download_salary_csv(month, year):
    import csv, tempfile, os
    cursor = mysql.connection.cursor()

    cursor.execute("SELECT emp_id, name, salary_per_day FROM employees")
    employees = cursor.fetchall()

    headers = [['EMP ID', 'Employee Name', 'Present Days', 'Salary', 'Bonus', 'Advance', 'Pending ','Outstanding'],['','','','','','','','']]
    records = []
    

    # Totals initialization
    total_salary = 0
    total_bonus = 0
    total_advance = 0
    total_pending = 0
    total_outstanding=0

    for emp_id, name, salary_per_day in employees:
        table_name = f"attendance_{emp_id}"

        try:
            cursor.execute(f"""
                SELECT COUNT(*) FROM {table_name}
                WHERE status = 'full' AND MONTH(date) = %s AND YEAR(date) = %s
            """, (month, year))
            present_days = cursor.fetchone()[0]
        except:
            present_days = 0

        # Bonus
        try:
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0) FROM bonuses
                WHERE emp_id = %s AND MONTH(date) = %s AND YEAR(date) = %s
            """, (emp_id, month, year))
            bonus = cursor.fetchone()[0]
        except:
            bonus = 0

        # Advance
        try:
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0) FROM advance
                WHERE emp_id = %s AND MONTH(date) = %s AND YEAR(date) = %s
            """, (emp_id, month, year))
            advance = cursor.fetchone()[0]
        except:
            advance = 0
        
        outstanding = 0

        # Salary and pending
        salary = salary_per_day * present_days
        pending = (salary + bonus) - advance
        if pending<0:
            
            outstanding = abs(pending)
        print(outstanding)

        # Accumulate totals
        total_salary += salary
        total_bonus += bonus
        total_advance += advance
        total_pending += pending
        total_outstanding+=outstanding

        records.append([emp_id, name, present_days, salary, bonus, advance, pending,outstanding])

    # Add total row
    records.append(['','','','','','','','',''])
    records.append(["Total", "", "", total_salary, total_bonus, total_advance, total_pending,total_outstanding])

    # Save to temp CSV
    filename = f"salary_summary_{month}_{year}.csv"
    temp_path = os.path.join(tempfile.gettempdir(), filename)
    a=[]
    c=int(total_salary+total_bonus)
    b=["","SALARY",c,"","","","",""]

    with open(temp_path, mode='w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(headers)
        writer.writerows(records)
        
        writer.writerow(a)
        writer.writerow(b)

    return send_file(temp_path, as_attachment=True)

from collections import defaultdict

@app.route('/download_advance_csv')
def download_advance_csv():
    cursor = mysql.connection.cursor()
    cursor.execute("SELECT employee_name, amount, date FROM advance")
    records = cursor.fetchall()

    # Step 1: Create CSV in memory
    si = StringIO()
    writer = csv.writer(si)

    # Step 2: Write Headers
    writer.writerow(['Employee Name', 'Advance Amount', 'Date'])
    writer.writerow(["", "", ""])  # Spacer

    total_advance = 0
    emp_totals = defaultdict(float)

    # Step 3: Write All Records and calculate totals
    for name, amount, date in records:
        total_advance += amount
        emp_totals[name] += amount
        writer.writerow([name, amount, date])

    # Spacer
    writer.writerow(["", "", ""])

    # Step 4: Overall total
    writer.writerow(["", "Total Advance Given", total_advance])
    writer.writerow(["", "", ""])  # Spacer

    # Step 5: Per-employee total summary
    writer.writerow(["Employee Name", "Total Advance"])
    for name, total in emp_totals.items():
        writer.writerow([name, total])

    output = si.getvalue()
    si.close()

    return Response(
        output,
        mimetype='text/csv',
        headers={"Content-Disposition": "attachment;filename=advance_records.csv"}
    )

@app.route('/download_bonus_csv')
def download_bonus_csv():
    cursor = mysql.connection.cursor()
    cursor.execute("SELECT employee_name, amount, date FROM bonuses")
    records = cursor.fetchall()

    # Step 1: Create CSV in memory
    si = StringIO()
    writer = csv.writer(si)

    # Step 2: Write Headers
    writer.writerow(['Employee Name', 'Bonus Amount', 'Date'])
    writer.writerow(["", "", ""])  # Spacer

    total_bonus = 0
    emp_totals = defaultdict(float)

    # Step 3: Write All Records and calculate totals
    for name, amount, date in records:
        total_bonus += amount
        emp_totals[name] += amount
        writer.writerow([name, amount, date])

    # Spacer
    writer.writerow(["", "", ""])

    # Step 4: Overall total
    writer.writerow(["", "Total Bonus Given", total_bonus])
    writer.writerow(["", "", ""])  # Spacer

    # Step 5: Per-employee total summary
    writer.writerow(["Employee Name", "Total Bonus"])
    for name, total in emp_totals.items():
        writer.writerow([name, total])

    output = si.getvalue()
    si.close()

    return Response(
        output,
        mimetype='text/csv',
        headers={"Content-Disposition": "attachment;filename=bonus_records.csv"}
    )

from flask import send_file, request
from io import BytesIO, StringIO
import calendar
import csv
from datetime import datetime
import MySQLdb.cursors

from flask import send_file, request
from io import BytesIO
import calendar
from datetime import datetime
import MySQLdb.cursors
from openpyxl import Workbook
from openpyxl.styles import PatternFill

@app.route('/download_calendar_csv')
def download_calendar_csv():
    try:
        month = int(request.args.get("month"))
        year = int(request.args.get("year"))
    except (TypeError, ValueError):
        return "Invalid or missing month/year", 400

    days_in_month = calendar.monthrange(year, month)[1]

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("SELECT emp_id, name FROM employees")
    employees = cursor.fetchall()

    # Create an Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month}-{year}"

    # Define cell fill colors
    full_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")   # Green
    half_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")   # Orange
    absent_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # Red

    # Header row
    headers = ["Employee Name"] + [str(day) for day in range(1, days_in_month + 1)] +['Total Days']
    ws.append(headers)
    ws.append([''])

    # Add attendance data
    for emp in employees:
        emp_id = emp['emp_id']
        name = emp['name']
        table_name = f"attendance_{emp_id}"

        row_index = ws.max_row + 1
        ws.cell(row=row_index, column=1, value=name)

        total_days_present = 0  # Track total days present for this employee

        for day in range(1, days_in_month + 1):
            date_str = f"{year}-{month:02d}-{day:02d}"
            cursor.execute(f"SELECT status FROM {table_name} WHERE date = %s", (date_str,))
            result = cursor.fetchone()

            col_index = day + 1  # column 1 is name
            cell = ws.cell(row=row_index, column=col_index)

            if not result:
                cell.value = ""
            elif result['status'] == 'full':
                cell.value = ""
                cell.fill = full_fill
                total_days_present += 1
            elif result['status'] == 'half':
                cell.value = ""
                cell.fill = half_fill
                total_days_present += 0.5
            elif result['status'] == 'absent':
                cell.value = ""
                cell.fill = absent_fill
            else:
                cell.value = ""

        # Set total present days in the last column
        total_col_index = days_in_month + 2
        ws.cell(row=row_index, column=total_col_index, value=total_days_present)
        ws.append([''])

    # Save to a BytesIO buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"attendance_{year}_{month:02d}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


    
if __name__ == '__main__':
    app.run(host="0.0.0.0",port=8080,debug=False)
    

